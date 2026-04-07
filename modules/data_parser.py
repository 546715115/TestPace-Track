"""
Excel 数据解析模块
处理合并单元格、字段合并、日期和进度标准化
"""
import openpyxl
import re
from datetime import datetime
from typing import List, Dict, Any, Tuple

# 字段类型定义
YES_NO_FIELDS = {'是否变更接口', '是否涉及资料', '是否涉及性能、过载', '是否涉及可靠性'}
PERSONNEL_FIELDS = {'测试人员', '开发人员', 'TSE', '业务团队'}
DATE_FIELDS = {'计划转测时间', '资料转测时间', '需求串讲/设计完成日期'}

# 进度列名（兼容旧版和新版Excel）
PROGRESS_FIELD = '测试进度（%）'
PROGRESS_FIELD_ALT = '测试进度'


def get_progress(row: Dict) -> Any:
    """获取进度值，兼容新旧列名"""
    return row.get(PROGRESS_FIELD) or row.get(PROGRESS_FIELD_ALT) or 0


def fix_encoding(value):
    """修复 Excel 读取时的编码问题

    Excel 文件使用 GBK/CP936 编码保存字符串，但 openpyxl 的 sharedStrings.xml
    读取时默认按 UTF-8 解码，导致 GBK 字节被误解析为 UTF-8 产生乱码。

    只对字符串值进行编码修复，非字符串值（int、float、bool）保持原样返回。
    """
    # 非字符串值保持原样（不包括 _row_span 等元数据）
    if value is None:
        return ''
    if isinstance(value, (int, float, bool)):
        return value
    if not isinstance(value, str):
        return value

    # 如果字符串已经是正确的汉字，直接返回
    if re.search(r'[\u4e00-\u9fff]', value):
        return value

    # 如果字符串是纯 ASCII（不包含可能出问题的字节），直接返回
    try:
        value.encode('ascii')
        return value
    except UnicodeEncodeError:
        pass

    # 检测是否包含 UTF-8 高位延续字节的不当序列
    # 0x80-0xBF 是 UTF-8 的延续字节，如果出现在字符串中说明可能被误读了
    # 通过检查是否包含非 ASCII 高位字符（而不是标准 UTF-8 合法字符）来判断
    try:
        # 尝试用 GBK 解码
        # 先把字符串当作 ISO-8859-1 编码的字节串（保留原始字节），再用 GBK 解码
        byte_data = value.encode('iso-8859-1', errors='ignore')
        decoded = byte_data.decode('gbk', errors='ignore')
        # 检查解码后是否包含正确的汉字，如果包含说明原来是 GBK 编码
        if re.search(r'[\u4e00-\u9fff]', decoded):
            return decoded
    except:
        pass

    return value


def fix_row_encoding(row: Dict) -> Dict:
    """修复一行数据的编码问题"""
    return {k: fix_encoding(v) for k, v in row.items()}


def parse_date(value) -> str:
    """解析多种日期格式，统一为 YYYY/MM/DD"""
    if value is None or str(value).strip() == '':
        return ''

    val = str(value).strip()

    # 检查是否包含日期分隔符，如果完全没有则直接返回原值
    if '/' not in val and '.' not in val:
        return str(value)

    # 已是目标格式
    if re.match(r'\d{4}/\d{1,2}/\d{1,2}', val):
        parts = val.split('/')
        try:
            return f"{int(parts[0]):04d}/{int(parts[1]):02d}/{int(parts[2]):02d}"
        except ValueError:
            return str(value)

    # 点分隔符
    if re.match(r'\d{4}\.\d{1,2}\.\d{1,2}', val):
        parts = val.split('.')
        try:
            return f"{int(parts[0]):04d}/{int(parts[1]):02d}/{int(parts[2]):02d}"
        except ValueError:
            return str(value)

    # 短格式（假设当前年份）
    if re.match(r'\d{1,2}/\d{1,2}', val):
        parts = val.split('/')
        try:
            year = datetime.now().year
            return f"{year}/{int(parts[0]):02d}/{int(parts[1]):02d}"
        except ValueError:
            return str(value)

    if re.match(r'\d{1,2}\.\d{1,2}', val):
        parts = val.split('.')
        try:
            year = datetime.now().year
            return f"{year}/{int(parts[0]):02d}/{int(parts[1]):02d}"
        except ValueError:
            return str(value)

    return str(value)


def normalize_progress(value) -> int:
    """标准化进度为整数百分比"""
    if value is None or str(value).strip() == '':
        return 0

    val = str(value).strip()

    if val == '已完成':
        return 100

    val = val.replace('%', '')

    try:
        num = float(val)
        # 如果原值是小数（如0.8表示80%），乘以100转为整数
        if num < 1 and num > 0:
            return int(num * 100)
        # 如果是整数（如80表示80%），直接返回
        if num == int(num):
            return int(num)
        else:
            return int(num)
    except:
        return 0


def merge_yes_no_field(values: list) -> str:
    """合并是否类字段：任意'是'则为'是'，有'否'则返回'否'，全空返回空"""
    has_yes = False
    has_no = False
    has_value = False
    for v in values:
        if v is None or str(v).strip() == '':
            continue
        has_value = True
        if v == '是':
            has_yes = True
        elif v == '否':
            has_no = True
    if has_yes:
        return '是'
    if has_no:
        return '否'
    if has_value:
        return '否'  # 有值但不是是/否，返回否
    return ''  # 全空


def merge_personnel_field(values: list) -> str:
    """合并人员类字段：去重拼接"""
    non_null = [v for v in values if v and str(v).strip() and str(v) != '/']
    unique = list(dict.fromkeys(non_null))  # 保持顺序，去重
    return ', '.join(unique) if unique else ''


def merge_value_field(values: list):
    """合并数值类字段：取第一个非空值"""
    for v in values:
        if v is not None and str(v).strip():
            return v
    return ''


def merge_progress_field(values: list):
    """合并进度类字段：取最小值"""
    progress_values = []
    for v in values:
        if v is not None and str(v).strip():
            progress_values.append(normalize_progress(v))
    return min(progress_values) if progress_values else 0


class ExcelReader:
    """Excel 读取器，解析合并单元格"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.current_sheet = None
        self._load()

    def _load(self):
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

    def get_sheet_names(self) -> List[str]:
        return self.workbook.sheetnames

    def get_headers(self, sheet_name: str) -> List[str]:
        ws = self.workbook[sheet_name]
        return [fix_encoding(cell.value) for cell in ws[1]]

    def load_sheet(self, sheet_name: str):
        self.current_sheet = sheet_name

    def get_merged_ranges(self) -> list:
        """获取所有合并单元格区域"""
        ws = self.workbook[self.current_sheet]
        return list(ws.merged_cells.ranges)

    def get_requirement_groups(self) -> list:
        """
        基于 Column A（特性分类）合并单元格检测需求组
        返回: [{'rows': [2,3,4,5], 'is_merged': True}, ...]
        """
        ws = self.workbook[self.current_sheet]
        groups = []
        merged_ranges = list(ws.merged_cells.ranges)

        # 找出 Column A 的合并区域
        col_a_merges = []
        for mr in merged_ranges:
            if mr.min_col == 1 and mr.max_col == 1:
                col_a_merges.append({
                    'min_row': mr.min_row,
                    'max_row': mr.max_row
                })

        # 添加合并组
        for mr in sorted(col_a_merges, key=lambda x: x['min_row']):
            groups.append({
                'rows': list(range(mr['min_row'], mr['max_row'] + 1)),
                'is_merged': True
            })

        # 找出被合并的行
        merged_rows = set()
        for mr in col_a_merges:
            for r in range(mr['min_row'], mr['max_row'] + 1):
                merged_rows.add(r)

        # 添加非合并的单行
        for row_idx in range(2, ws.max_row + 1):
            if row_idx not in merged_rows:
                groups.append({
                    'rows': [row_idx],
                    'is_merged': False
                })

        return groups

    def merge_group(self, group: dict) -> dict:
        """
        合并需求组为单行数据
        根据字段类型应用不同的合并逻辑
        """
        ws = self.workbook[self.current_sheet]
        headers = self.get_headers(self.current_sheet)
        result = {}

        for col_idx, header in enumerate(headers, start=1):
            values = []

            for row_idx in group['rows']:
                cell = ws.cell(row=row_idx, column=col_idx)
                values.append(fix_encoding(cell.value))

            # 根据字段类型应用合并逻辑
            if header in YES_NO_FIELDS:
                result[header] = merge_yes_no_field(values)
            elif header in PERSONNEL_FIELDS:
                result[header] = merge_personnel_field(values)
            elif any(kw in str(header) for kw in ['进度', '进度（%）']):
                result[header] = merge_progress_field(values)
            elif any(kw in str(header) for kw in ['时间', '日期']):
                result[header] = parse_date(merge_value_field(values))
            else:
                result[header] = merge_value_field(values)

        # 保留元数据
        result['_rows'] = group['rows']
        result['_is_merged'] = group['is_merged']
        return result

    def get_all_requirements(self) -> Tuple[List[Dict], List[Dict]]:
        """
        获取所有解析后的需求和分组信息
        返回: (merged_requirements, groups)
        """
        groups = self.get_requirement_groups()
        merged = [self.merge_group(g) for g in groups]
        return merged, groups

    def get_raw_rows(self) -> Tuple[List[Dict], List[Dict]]:
        """
        获取原始所有行数据和分组信息（用于前端rowspan合并）
        返回: (raw_rows, groups)
        - raw_rows: 每行原始数据，合并单元格的值会复制到同组所有行
        - groups: 合并组信息，用于计算rowspan
        """
        ws = self.workbook[self.current_sheet]
        headers = self.get_headers(self.current_sheet)
        groups = self.get_requirement_groups()

        print(f'[DEBUG get_raw_rows] Sheet: {self.current_sheet}, ws.max_row: {ws.max_row}, headers count: {len(headers)}, groups count: {len(groups)}')

        # 构建 group -> 第一行数据的映射（用于复制合并单元格的值）
        group_first_row_data = {}
        for idx, group in enumerate(groups):
            if group['rows']:
                first_row_idx = group['rows'][0]
                first_row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=first_row_idx, column=col_idx)
                    first_row_data[header] = cell.value
                group_first_row_data[idx] = first_row_data

        # 构建 group -> row_indices 的映射
        group_of_row = {}
        for idx, group in enumerate(groups):
            for row_idx in group['rows']:
                group_of_row[row_idx] = idx

        # 获取所有原始行数据
        raw_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data[header] = cell.value

            # 添加元数据：属于哪个组、是否是合并组的第一行
            group_idx = group_of_row.get(row_idx)
            if group_idx is not None:
                group = groups[group_idx]
                row_data['_group_idx'] = group_idx
                row_data['_is_first_in_group'] = (row_idx == group['rows'][0])
                row_data['_is_merged'] = group['is_merged']
                row_data['_row_span'] = len(group['rows']) if group['is_merged'] else 1

                # 如果是合并组且不是第一行，复制第一行的值（用于合并显示）
                if not row_data['_is_first_in_group'] and group_first_row_data.get(group_idx):
                    for header in headers:
                        if row_data.get(header) is None and group_first_row_data[group_idx].get(header) is not None:
                            row_data[header] = group_first_row_data[group_idx][header]
            else:
                row_data['_group_idx'] = -1
                row_data['_is_first_in_group'] = True
                row_data['_is_merged'] = False
                row_data['_row_span'] = 1

            raw_rows.append(fix_row_encoding(row_data))

        return raw_rows, groups


if __name__ == '__main__':
    # 测试代码
    import os
    test_file = os.path.join(os.path.dirname(__file__), '..', 'test_read_copy.xlsx')
    if os.path.exists(test_file):
        reader = ExcelReader(test_file)
        print("Sheets:", reader.get_sheet_names())

        reader.load_sheet('0330需求列表')
        groups = reader.get_requirement_groups()
        print(f"Found {len(groups)} groups")

        if groups:
            merged = reader.merge_group(groups[0])
            print("First merged row keys:", list(merged.keys())[:5])
