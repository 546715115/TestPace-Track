# -*- coding: utf-8 -*-
import openpyxl
import random

wb = openpyxl.load_workbook('测试进度模板_v2.xlsx')
ws = wb['0330需求列表']

headers = [cell.value for cell in ws[1]]

# 列索引
col_map = {}
for i, h in enumerate(headers):
    if h:
        col_map[h] = i + 1

# 定义池
teams = ['CES Agent', 'CES', 'CES项目', 'DI工具', '报表系统', '权限中心', '消息中心', '订单系统']
developers = ['张三', '李四', '王五', '赵六', '钱七', '孙八', '周九', '吴十']
tse_list = ['TSE-A', 'TSE-B', 'TSE-C', 'TSE-D']
testers = ['测试A', '测试B', '测试C', '测试D', '测试E', '测试F']
test_types = ['功能测试', '集成测试', '系统测试', '回归测试']
progress_designs = ['已完成', '进行中', '未开始', '部分完成']
yes_no = ['是', '否', '/']
databases = ['MySQL', 'MySQL', 'Cassandra', 'influxDB', 'MySQL+Cassandra', '无', '/']
progress_details = ['正常推进', '有风险', '受阻', '等待上游', '资源不足']
notes = ['无', '正常', '需关注', '优先级调整', '/', '高风险', '关键路径']

# 合并组配置
groups = [
    (2, 5, {}),
    (10, 13, {}),
    (19, 22, {}),
    (27, 30, {}),
    (35, 36, {}),
]

random.seed(42)

# 为每个合并组生成组内统一的基础数据
for i, (min_row, max_row, group_data) in enumerate(groups):
    group_data['team'] = random.choice(teams)
    group_data['dev'] = random.choice(developers)
    group_data['tse'] = random.choice(tse_list)
    group_data['tester'] = random.choice(testers)
    group_data['test_type'] = random.choice(test_types)
    group_data['has_interface_change'] = random.choice(['是', '否', '否', '否'])
    group_data['has_doc'] = random.choice(['是', '否', '否', '否'])
    group_data['has_perf'] = random.choice(['是', '否', '否', '否', '否'])
    group_data['has_reliability'] = random.choice(['是', '否', '否', '否', '否'])
    group_data['database'] = random.choice(databases)
    group_data['test_progress'] = round(random.choice([0.3, 0.5, 0.7, 0.8, 0.9]), 1)
    group_data['test_case_count'] = random.choice([30, 50, 80, 100, 120])
    group_data['progress_design'] = random.choice(progress_designs)
    group_data['progress_review'] = random.choice(['0', '0.3', '0.5', '0.7', '1'])
    group_data['issue_count'] = random.randint(0, 15)

# 非合并行
non_merged_rows = [6, 7, 8, 9, 14, 15, 16, 17, 18, 23, 24, 25, 26, 31, 32, 33, 34, 37]

# 填充合并组数据
for min_row, max_row, group_data in groups:
    for row_idx in range(min_row, max_row + 1):
        ws.cell(row=row_idx, column=col_map['业务团队']).value = group_data['team']
        ws.cell(row=row_idx, column=col_map['开发人员']).value = group_data['dev']
        ws.cell(row=row_idx, column=col_map['TSE']).value = group_data['tse']
        ws.cell(row=row_idx, column=col_map['测试人员']).value = group_data['tester']
        ws.cell(row=row_idx, column=col_map['测试类型']).value = group_data['test_type']
        ws.cell(row=row_idx, column=col_map['上线范围']).value = '本迭代'
        ws.cell(row=row_idx, column=col_map['涉及数据底座（MySQL/Cassandra/influxDB）']).value = group_data['database']
        ws.cell(row=row_idx, column=col_map['是否变更接口']).value = group_data['has_interface_change']
        ws.cell(row=row_idx, column=col_map['是否涉及资料']).value = group_data['has_doc']
        ws.cell(row=row_idx, column=col_map['是否涉及性能、过载']).value = group_data['has_perf']
        ws.cell(row=row_idx, column=col_map['是否涉及可靠性']).value = group_data['has_reliability']

        base_progress = group_data['test_progress']
        variance = random.uniform(-0.1, 0.1)
        ws.cell(row=row_idx, column=col_map['测试进度']).value = round(base_progress + variance, 1)

        base_cases = group_data['test_case_count']
        case_variance = random.randint(-20, 20)
        ws.cell(row=row_idx, column=col_map['用例数']).value = max(10, base_cases + case_variance)

        ws.cell(row=row_idx, column=col_map['串讲和测试设计进度']).value = random.choice(progress_designs)
        ws.cell(row=row_idx, column=col_map['反串讲进度（%）']).value = group_data['progress_review']
        ws.cell(row=row_idx, column=col_map['问题单数量']).value = group_data['issue_count']
        ws.cell(row=row_idx, column=col_map['进展详情']).value = random.choice(progress_details)

        plan_dates = ['2024/03/15', '2024/03/20', '2024/03/25', '2024/04/01', '2024/04/05', '2024/04/10', '待定']
        ws.cell(row=row_idx, column=col_map['计划转测时间（格式:2024/01/01）']).value = random.choice(plan_dates)

        if group_data['has_doc'] == '是':
            doc_dates = ['2024/03/20', '2024/03/25', '2024/04/01', '2024/04/05']
            ws.cell(row=row_idx, column=col_map['资料转测时间']).value = random.choice(doc_dates)
        else:
            ws.cell(row=row_idx, column=col_map['资料转测时间']).value = '/'

        qualities = ['pass', 'fail', 'pass/fail', '/']
        ws.cell(row=row_idx, column=col_map['自验质量（自验pass，测试fail）']).value = random.choice(qualities)
        ws.cell(row=row_idx, column=col_map['备注']).value = random.choice(notes)

# 填充非合并行
for row_idx in non_merged_rows:
    ws.cell(row=row_idx, column=col_map['业务团队']).value = random.choice(teams)
    ws.cell(row=row_idx, column=col_map['开发人员']).value = random.choice(developers)
    ws.cell(row=row_idx, column=col_map['TSE']).value = random.choice(tse_list)
    ws.cell(row=row_idx, column=col_map['测试人员']).value = random.choice(testers)
    ws.cell(row=row_idx, column=col_map['上线范围']).value = '本迭代'
    ws.cell(row=row_idx, column=col_map['串讲和测试设计进度']).value = random.choice(progress_designs)
    ws.cell(row=row_idx, column=col_map['反串讲进度（%）']).value = random.choice(['0', '0.3', '0.5', '0.7', '1', '/'])
    ws.cell(row=row_idx, column=col_map['用例数']).value = random.choice([30, 50, 80, 100, 120, 150, 200])
    ws.cell(row=row_idx, column=col_map['计划转测时间（格式:2024/01/01）']).value = random.choice(['2024/03/15', '2024/03/20', '2024/03/25', '2024/04/01', '2024/04/05', '2024/04/10', '待定', '已转'])
    ws.cell(row=row_idx, column=col_map['测试类型']).value = random.choice(test_types)
    ws.cell(row=row_idx, column=col_map['测试进度']).value = round(random.uniform(0.3, 1.0), 1)
    ws.cell(row=row_idx, column=col_map['自验质量（自验pass，测试fail）']).value = random.choice(['pass', 'fail', 'pass/fail', '/'])
    ws.cell(row=row_idx, column=col_map['问题单数量']).value = random.randint(0, 20)
    ws.cell(row=row_idx, column=col_map['进展详情']).value = random.choice(progress_details)
    ws.cell(row=row_idx, column=col_map['是否变更接口']).value = random.choice(yes_no)
    ws.cell(row=row_idx, column=col_map['是否涉及资料']).value = random.choice(yes_no)
    ws.cell(row=row_idx, column=col_map['资料转测时间']).value = random.choice(['2024/03/20', '2024/03/25', '2024/04/01', '2024/04/05', '/', '待定', '已转'])
    ws.cell(row=row_idx, column=col_map['是否涉及性能、过载']).value = random.choice(yes_no)
    ws.cell(row=row_idx, column=col_map['是否涉及可靠性']).value = random.choice(yes_no)
    ws.cell(row=row_idx, column=col_map['涉及数据底座（MySQL/Cassandra/influxDB）']).value = random.choice(databases)
    ws.cell(row=row_idx, column=col_map['备注']).value = random.choice(notes)

wb.save('测试进度模板_v3.xlsx')
print("已生成: 测试进度模板_v3.xlsx")
print("\n合并组数据预览:")
for i, (min_row, max_row, group_data) in enumerate(groups):
    print(f"  组{i+1} 行{min_row}-{max_row}: 团队={group_data['team']}, 测试人员={group_data['tester']}, 变更={group_data['has_interface_change']}, 涉及资料={group_data['has_doc']}")